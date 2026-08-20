#!/usr/bin/env python3
"""Devolve para o banco o que foi editado na planilha gerada por exportar_planilha.py.

Só escreve o que mudou de verdade e só nas colunas editáveis — as demais são reescritas na
próxima exportação, então mexer nelas não teria efeito. A chave é sempre `codigo`.

  python scripts/import_edicoes.py --xlsx duck-editavel.xlsx --out db/seed_05_edicoes.sql
  psql -d duck -f db/seed_05_edicoes.sql
"""
import argparse
from pathlib import Path
import openpyxl


def s(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return "NULL"
    return "'" + str(v).strip().replace("'", "''") + "'"


def n(v):
    if v in (None, ""):
        return "NULL"
    try:
        return repr(round(float(str(v).replace("R$", "").replace(".", "").replace(",", ".")
                               if isinstance(v, str) else v), 2))
    except (TypeError, ValueError):
        return "NULL"


def linhas(ws):
    it = ws.iter_rows(values_only=True)
    cab = [str(c).strip() if c else "" for c in next(it)]
    for r in it:
        if r and r[0]:
            yield dict(zip(cab, r))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True)
    ap.add_argument("--out", default="db/seed_05_edicoes.sql")
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.xlsx, data_only=True)
    sql = ["-- GERADO POR scripts/import_edicoes.py a partir de " + Path(a.xlsx).name,
           "-- Camada mais recente: aplicar DEPOIS de todos os outros seeds.", ""]
    contagem = {}

    if "Itens" in wb.sheetnames:
        c = 0
        for r in linhas(wb["Itens"]):
            partes = []
            if r.get("valor_reposicao") not in (None, ""):
                partes.append(f"valor_reposicao = {n(r['valor_reposicao'])}")
            conf = str(r.get("confirmado") or "").strip().upper()
            if conf in ("S", "SIM", "TRUE", "1"):
                partes.append("valor_reposicao_confirmado = true")
            if r.get("numero_serie") not in (None, ""):
                partes.append(f"numero_serie = {s(r['numero_serie'])}")
            if r.get("observacoes") not in (None, ""):
                partes.append(f"observacoes = {s(r['observacoes'])}")
            if partes:
                sql.append(f"UPDATE asset SET {', '.join(partes)} WHERE codigo = {s(r['codigo'])};")
                c += 1
            case = r.get("dentro_do_case")
            if case not in (None, ""):
                sql.append(f"UPDATE asset a SET container_id = c.id FROM asset c "
                           f"WHERE c.codigo = {s(case)} AND a.codigo = {s(r['codigo'])};")
        contagem["itens"] = c

    if "Sublocados" in wb.sheetnames:
        c = 0
        for r in linhas(wb["Sublocados"]):
            if r.get("fornecedor") not in (None, ""):
                # cria o fornecedor se ainda não existir, depois liga ao item
                sql.append(f"INSERT INTO company (nome, tipo) VALUES ({s(r['fornecedor'])}, "
                           f"'fornecedor') ON CONFLICT DO NOTHING;")
                sql.append(f"UPDATE asset a SET fornecedor_id = f.id FROM company f "
                           f"WHERE f.nome = {s(r['fornecedor'])} AND a.codigo = {s(r['codigo'])};")
            partes = []
            if r.get("custo_diaria") not in (None, ""):
                partes.append(f"custo_diaria = {n(r['custo_diaria'])}")
            if r.get("valor_diaria") not in (None, ""):
                partes.append(f"valor_diaria = {n(r['valor_diaria'])}, requer_cotacao = false")
            if partes:
                sql.append(f"UPDATE asset SET {', '.join(partes)} WHERE codigo = {s(r['codigo'])};")
                c += 1
        contagem["sublocados"] = c

    if "Precos_locacao" in wb.sheetnames:
        c = 0
        for r in linhas(wb["Precos_locacao"]):
            partes = [f"{col} = {n(r[col])}" for col in
                      ("valor_diaria", "valor_semanal", "valor_mensal")
                      if r.get(col) not in (None, "")]
            if partes:
                sql.append(f"UPDATE asset SET {', '.join(partes)} WHERE codigo = {s(r['codigo'])};")
                c += 1
        contagem["preços de locação"] = c

    if "Precos_servico" in wb.sheetnames:
        c = 0
        for r in linhas(wb["Precos_servico"]):
            ativo = str(r.get("ativo") or "true").strip().lower() not in ("f", "false", "n", "não", "nao", "0")
            sql.append(
                "INSERT INTO price_list (codigo, descricao, unidade, valor, categoria, ativo) VALUES ("
                f"{s(r['codigo'])}, {s(r.get('descricao'))}, {s(r.get('unidade'))}, "
                f"{n(r.get('valor'))}, {s(r.get('categoria'))}, {str(ativo).lower()}) "
                "ON CONFLICT (codigo) DO UPDATE SET descricao = EXCLUDED.descricao, "
                "unidade = EXCLUDED.unidade, valor = EXCLUDED.valor, "
                "categoria = EXCLUDED.categoria, ativo = EXCLUDED.ativo;")
            c += 1
        contagem["preços de serviço"] = c

    Path(a.out).write_text("\n".join(sql) + "\n", encoding="utf-8")
    print(f"-> {a.out}")
    for k, v in contagem.items():
        print(f"   {v} {k}")


if __name__ == "__main__":
    main()
