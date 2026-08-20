#!/usr/bin/env python3
"""Gera a planilha editável a partir do banco.

O que só você sabe — valor de reposição, número de série, custo de sublocação, preço de serviço —
não deve morar em SQL. Mora aqui, você edita no Excel/Numbers, e `import_edicoes.py` devolve para
o banco. Nenhum dado é perdido no caminho: a chave é sempre o código do item.

  python scripts/exportar_planilha.py --db duck --out duck-editavel.xlsx
"""
import argparse, csv, io, subprocess, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CINZA = PatternFill("solid", fgColor="1B2725")
EDIT = PatternFill("solid", fgColor="FFF6DA")     # amarelo claro = você pode editar
CAB = Font(color="FFFFFF", bold=True)

ABAS = {
    "Itens": ("""
        SELECT codigo, nome, categoria, marca,
               valor_aquisicao, valor_reposicao,
               CASE WHEN valor_reposicao_confirmado THEN 'S' ELSE 'N' END AS confirmado,
               numero_serie,
               (SELECT c.codigo FROM asset c WHERE c.id = a.container_id) AS dentro_do_case,
               observacoes
          FROM asset a WHERE proprietario = 'proprio' ORDER BY codigo""",
        ["valor_reposicao", "confirmado", "numero_serie", "dentro_do_case", "observacoes"]),

    "Sublocados": ("""
        SELECT codigo, nome, categoria, marca,
               (SELECT f.nome FROM company f WHERE f.id = a.fornecedor_id) AS fornecedor,
               custo_diaria, valor_diaria, observacoes
          FROM asset a WHERE proprietario = 'sublocado' ORDER BY codigo""",
        ["fornecedor", "custo_diaria", "valor_diaria", "observacoes"]),

    "Precos_servico": ("""
        SELECT codigo, descricao, unidade, valor, categoria, ativo
          FROM price_list ORDER BY categoria, codigo""",
        ["descricao", "unidade", "valor", "categoria", "ativo"]),

    "Precos_locacao": ("""
        SELECT codigo, nome, categoria,
               valor_aquisicao, valor_diaria, valor_semanal, valor_mensal,
               round(100 * valor_diaria / NULLIF(valor_aquisicao, 0), 2) AS pct_diaria
          FROM asset WHERE proprietario = 'proprio' ORDER BY codigo""",
        ["valor_diaria", "valor_semanal", "valor_mensal"]),
}

INSTRUCOES = [
    ("Como usar esta planilha", True),
    ("", False),
    ("Colunas em AMARELO são suas: edite à vontade.", False),
    ("Colunas em cinza vêm do sistema — não edite, elas são reescritas na próxima exportação.", False),
    ("A coluna 'codigo' é a chave. Não altere e não apague linhas.", False),
    ("", False),
    ("Itens", True),
    ("valor_reposicao: quanto custa REPOR o item hoje. É o número que vai no termo de", False),
    ("  responsabilidade — não é o que você pagou. Comece pelos ~30 mais caros.", False),
    ("confirmado: escreva S quando o valor for conferido por você. Enquanto estiver N,", False),
    ("  o app avisa em toda saída que o valor é estimativa.", False),
    ("numero_serie: capture na mesma passada em que colar as etiquetas.", False),
    ("  Sem série, seguro e B.O. ficam muito mais difíceis.", False),
    ("dentro_do_case: código do case onde o item mora (ex.: 0139). Deixe vazio se não mora em case.", False),
    ("", False),
    ("Sublocados", True),
    ("Itens que você aluga de terceiros. fornecedor + custo_diaria são o que falta para", False),
    ("  o sistema calcular margem e para o agente saber a quem ligar antes de fechar preço.", False),
    ("", False),
    ("Precos_locacao", True),
    ("Editar valor_diaria/semanal/mensal muda o preço daquele item específico.", False),
    ("pct_diaria é calculado (diária ÷ valor de compra) — serve de conferência:", False),
    ("  3,5% câmera principal · 3% premium · 2,5% médio · 2% case.", False),
    ("", False),
    ("Depois de editar", True),
    ("python scripts/import_edicoes.py --xlsx duck-editavel.xlsx --out db/seed_05_edicoes.sql", False),
    ("psql -d duck -f db/seed_05_edicoes.sql", False),
]


def consulta(db, sql):
    out = subprocess.run(["psql", "-d", db, "-v", "ON_ERROR_STOP=1", "--csv", "-c", " ".join(sql.split())],
                         capture_output=True, text=True)
    if out.returncode:
        sys.exit(f"psql falhou:\n{out.stderr}")
    return list(csv.reader(io.StringIO(out.stdout)))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--db", default="duck")
    ap.add_argument("--out", default="duck-editavel.xlsx")
    a = ap.parse_args()

    wb = Workbook(); wb.remove(wb.active)

    ws = wb.create_sheet("Instruções")
    for i, (txt, forte) in enumerate(INSTRUCOES, 1):
        c = ws.cell(i, 1, txt)
        if forte:
            c.font = Font(bold=True, size=12)
    ws.column_dimensions["A"].width = 100

    for nome, (sql, editaveis) in ABAS.items():
        linhas = consulta(a.db, sql)
        if not linhas:
            continue
        ws = wb.create_sheet(nome)
        cab = linhas[0]
        for j, h in enumerate(cab, 1):
            c = ws.cell(1, j, h); c.font = CAB; c.fill = CINZA
            c.alignment = Alignment(horizontal="center")
        for i, linha in enumerate(linhas[1:], 2):
            for j, v in enumerate(linha, 1):
                cel = ws.cell(i, j, _tipa(v, cab[j - 1]))
                if cab[j - 1] in editaveis:
                    cel.fill = EDIT
        for j, h in enumerate(cab, 1):
            largura = max(len(h), *(len(l[j - 1]) for l in linhas[1:])) + 3
            ws.column_dimensions[get_column_letter(j)].width = min(largura, 46)
        ws.freeze_panes = "B2"
        print(f"{nome}: {len(linhas)-1} linhas")

    wb.save(a.out)
    print(f"-> {a.out}")


# Colunas que são identificador, nunca número: "0118" precisa continuar sendo "0118".
# Converter para int aqui apaga o zero à esquerda e quebra o casamento na volta.
TEXTO = {"codigo", "numero_serie", "dentro_do_case", "unidade", "categoria"}


def _tipa(v, coluna=""):
    if v == "":
        return None
    if coluna in TEXTO:
        return v
    try:
        return float(v) if "." in v else int(v)
    except ValueError:
        return v


if __name__ == "__main__":
    main()
