#!/usr/bin/env python3
"""Importa o inventário do AssetTiger + a tabela de aluguel para o schema do CRM.

Entradas (xlsx exportados hoje):
  --assets   export do AssetTiger  (Asset Tag ID, Description, Brand, Purchase Date, Cost, Status)
  --precos   planilha de aluguel   (Asset Tag, Descrição, Marca, Valor de Compra, Categoria,
                                    % Diária, % Semanal, Diária, Semanal, Status)

Saídas:
  db/seed_inventario.sql   carga idempotente (ON CONFLICT) da tabela `asset`
  relatório de divergências no stdout

Regra de conflito: onde as duas planilhas discordam do valor de compra, vence a **planilha de
aluguel** (é ela que gera preço hoje) e a divergência é reportada para decisão humana.
"""
import argparse, json, re, sys, unicodedata
from pathlib import Path
import openpyxl

# Grafias divergentes encontradas nas duas planilhas -> forma canônica
MARCAS = {"smallrig": "SmallRig", "nisi": "NiSi", "zoom": "Zoom", "generico": "Genérico",
          "generica": "Genérico", "k&f concept": "K&F Concept", "dji": "DJI", "rode": "RODE",
          "skb": "SKB", "dewalt": "DeWalt", "inspired energy": "Inspired Energy"}

CONTAINER_RE = re.compile(r"\b(case|pelican|toughsystem|hard case|maleta)\b", re.I)


def norm_marca(m):
    m = (m or "").strip()
    return MARCAS.get(m.lower(), m)


def num(v):
    """Aceita 25000, '25,000.00' (US) e '25.000,00' (pt-BR)."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    t = str(v).strip().replace(" ", "")
    if "," in t and "." in t:            # o separador decimal é o último que aparece
        dec = max(t.rfind(","), t.rfind("."))
        t = t[:dec].replace(",", "").replace(".", "") + "." + t[dec + 1:]
    elif "," in t:                        # vírgula sozinha: decimal só se não for milhar
        t = t.replace(",", "") if len(t.split(",")[-1]) == 3 else t.replace(",", ".")
    try:
        return float(t)
    except ValueError:
        return None


def sqlstr(v):
    if v is None or v == "":
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


def ler(path, cols):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(c).strip() if c else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {}
    for key, nomes in cols.items():
        for n in nomes:
            if n in header:
                idx[key] = header.index(n)
                break
        else:
            sys.exit(f"ERRO: coluna {nomes} não encontrada em {path} (header: {header})")
    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[idx["tag"]]:
            continue
        d = {k: r[i] for k, i in idx.items()}
        d["tag"] = str(d["tag"]).strip().zfill(4)
        out[d["tag"]] = d
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--assets", required=True)
    ap.add_argument("--precos", required=True)
    ap.add_argument("--out", default="db/seed_inventario.sql")
    a = ap.parse_args()

    A = ler(a.assets, {"tag": ["Asset Tag ID", "Asset Tag"], "desc": ["Description", "Descrição"],
                       "marca": ["Brand", "Marca"], "data": ["Purchase Date"], "custo": ["Cost"],
                       "status": ["Status"]})
    P = ler(a.precos, {"tag": ["Asset Tag", "Asset Tag ID"], "desc": ["Descrição", "Description"],
                       "marca": ["Marca", "Brand"], "custo": ["Valor de Compra"],
                       "cat": ["Categoria"], "pct_d": ["% Diária"], "pct_s": ["% Semanal"],
                       "diaria": ["Diária (R$)"], "semanal": ["Semanal (R$)"]})

    divergencias, linhas = [], []
    for tag in sorted(set(A) | set(P)):
        ai, pi = A.get(tag), P.get(tag)
        if not ai:
            divergencias.append(("so_no_aluguel", tag, "", ""))
        if not pi:
            divergencias.append(("sem_preco", tag, ai["desc"], ""))

        # descrição: a do AssetTiger é mais completa (inclui a marca em vários itens)
        desc = (ai or pi)["desc"].strip()
        if ai and pi and desc.lower() != str(pi["desc"]).strip().lower():
            divergencias.append(("descricao", tag, desc, str(pi["desc"]).strip()))

        marca = norm_marca((ai or pi).get("marca"))
        if ai and pi and norm_marca(ai["marca"]) != norm_marca(pi["marca"]):
            divergencias.append(("marca", tag, ai["marca"], pi["marca"]))

        custo_a = num(ai["custo"]) if ai else None
        custo_p = num(pi["custo"]) if pi else None
        if custo_a is not None and custo_p is not None and abs(custo_a - custo_p) > 0.005:
            divergencias.append(("valor", tag, f"{custo_a:.2f}", f"{custo_p:.2f}"))
        custo = custo_p if custo_p is not None else custo_a   # aluguel vence

        data = str(ai["data"])[:10] if ai and ai["data"] else None
        if not data:
            divergencias.append(("sem_data", tag, desc, ""))

        cat = (pi or {}).get("cat") or "Sem categoria"
        diaria = num(pi["diaria"]) if pi else None
        semanal = num(pi["semanal"]) if pi else None
        e_container = bool(CONTAINER_RE.search(desc)) or cat == "Case/Proteção"

        linhas.append(
            f"  ({sqlstr(tag)}, {sqlstr(desc)}, {sqlstr(cat)}, {sqlstr(marca)}, "
            f"{custo if custo is not None else 'NULL'}, "
            f"{diaria if diaria is not None else 'NULL'}, "
            f"{semanal if semanal is not None else 'NULL'}, "
            f"{sqlstr(data)}::date, {str(e_container).lower()})")

    out = Path(a.out)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(
        "-- GERADO POR scripts/import_inventario.py — não editar à mão.\n"
        "-- Fonte: export do AssetTiger + planilha de aluguel.\n"
        "-- Idempotente: rodar de novo atualiza os campos, preserva ids e histórico.\n"
        "-- valor_reposicao fica NULL de propósito: é o valor que vai no termo de\n"
        "-- responsabilidade e precisa ser o custo de repor HOJE, não o de compra.\n\n"
        "INSERT INTO asset (codigo, nome, categoria, marca, valor_aquisicao, valor_diaria,\n"
        "                   valor_semanal, data_aquisicao, e_container, serializado, origem_import)\n"
        "VALUES\n" + ",\n".join(l[:-1] + ", true, 'assettiger:2026-08')" for l in linhas) + "\n"
        "ON CONFLICT (codigo) DO UPDATE SET\n"
        "  nome = EXCLUDED.nome, categoria = EXCLUDED.categoria, marca = EXCLUDED.marca,\n"
        "  valor_aquisicao = EXCLUDED.valor_aquisicao, valor_diaria = EXCLUDED.valor_diaria,\n"
        "  valor_semanal = EXCLUDED.valor_semanal,\n"
        "  data_aquisicao = EXCLUDED.data_aquisicao, e_container = EXCLUDED.e_container;\n",
        encoding="utf-8")

    print(f"{len(linhas)} itens -> {out}")
    if divergencias:
        print(f"\n{len(divergencias)} divergências (decisão humana):")
        for tipo, tag, x, y in divergencias:
            print(f"  [{tipo:<14}] {tag}  {x}" + (f"   ->   {y}" if y else ""))


if __name__ == "__main__":
    main()
